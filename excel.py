from scraper import retrieve_data, set_filters
from prompts import get_input
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def export_to_excel(car_heap, file_name="auto_dati.xlsx"):
    """Saglabā automašīnu sarakstu Excel failā"""
    if not car_heap or not car_heap.heap:
        print("Nav datu, ko saglabāt.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Auto dati"

    # Virsraksta rinda
    headers = ["Saite", "Īss apraksts", "Gads", "Tilpums (L)", "Nobraukums (tūkst. km)", "Cena (€)"]
    ws.append(headers)


    for car in car_heap.heap:
        row = [
            car.link,
            car.text,
            car.year,
            car.engine_size,
            car.mileage,
            car.price
        ]
        ws.append(row)

    # Formatē virsrakstus un kolonnas
    for col_num, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
        max_length = max(len(str(cell.value)) for cell in column_cells)
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = max(max_length + 2, 15)

        for cell in column_cells:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    # Formatē cenas kā valūtu
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = '#,## €'

    wb.save(file_name)
    print(f"Dati saglabāti failā: {file_name}")

if __name__ == "__main__":
    print("Sākam programmu...")


    data = get_input()
    print("Lietotāja dati saņemti:", data)

    # Kārtošanas kritērijs
    excel_filter = input("Izvēlieties Excel filtrus: gads, tilpums, nobraukums, cena -> ").strip()
    print(f"Izvēlētais filtrs: {excel_filter}")

  
    driver = webdriver.Chrome()
    print("Pārlūks palaists.")

    try:
        filters = set_filters(driver, data)
        print("Filtri uzstādīti:", filters)

        if filters:
            car_heap = retrieve_data(driver, excel_filter)
            print(f"Atrasti {len(car_heap.heap)} ieraksti.")
            export_to_excel(car_heap)
        else:
            print("Nav derīgu filtru. Dati netika iegūti.")
    except Exception as e:
        print("Kļūda izpildes laikā:", e)
    finally:
        driver.quit()
        print("Pārlūks aizvērts.")
